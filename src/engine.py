from openpyxl import load_workbook
import pandas as pd

def read_table1(file):

    wb = load_workbook(file, data_only=True)
    ws = wb["Résultats"]

    # récupérer la table
    table = ws.tables["Table1"]

    # plage de la table (ex: A10:F40)
    data = ws[table.ref]

    # convertir en liste
    rows = list(data)

    # header
    columns = [cell.value for cell in rows[0]]

    # data
    values = [
        [cell.value for cell in row]
        for row in rows[1:]
    ]

    df = pd.DataFrame(values, columns=columns)

    # nettoyage
    df.columns = df.columns.str.strip().str.lower()

    return df


# ----------------------
# SAFE CHECK (Résultats)
# ----------------------
def safe_check(df, crit):

    # Nettoyage colonnes
    df.columns = df.columns.str.strip().str.lower()

    # 🔍 trouver colonnes dynamiquement
    col_j = None
    col_r = None

    for col in df.columns:
        if "jaune" in col:
            col_j = col
        if "rouge" in col:
            col_r = col

    if col_j is None or col_r is None:
        raise ValueError(f"❌ Colonnes jaunes/rouges introuvables : {df.columns.tolist()}")

    # sécurisation valeurs
    df[col_j] = df[col_j].fillna(0)
    df[col_r] = df[col_r].fillna(0)

    # filtre ITEM
    if "item" not in df.columns:
        raise ValueError("❌ Colonne ITEM absente")

    row = df[df["item"] == crit]

    if row.empty:
        return 0

    return int((row[col_j].values[0] + row[col_r].values[0]) > 0)
    print(df.columns.tolist())


# ----------------------
# POSTURES MEC / VEC
# ----------------------
def compute_postures_mec(df):
    return {
        "membres_inf": safe_check(df, 3),
        "poignet": safe_check(df, 8),
        "epaule": safe_check(df, 9),
        "dos": safe_check(df, 10),
        "cervicales": safe_check(df, 11)
    }


def compute_postures_vec(df):
    return {
        "membres_inf": int(safe_check(df, 4) or safe_check(df, 1)),
        "poignet": safe_check(df, 5),
        "epaule": safe_check(df, 6),
        "dos": int(safe_check(df, 7) or safe_check(df, 3)),
        "cervicales": int(safe_check(df, 8) or safe_check(df, 2))
    }


# ----------------------
# CHARGES
# ----------------------
def extract_charges(file, is_mec):

    df = pd.read_excel(file, sheet_name="Evaluation ergonomique", header=None)

    if is_mec:
        mapping = {
            "3_6_KG": (27, 9),
            "6_9_KG": (28, 9),
            "9_12_KG": (29, 9),
            "12+_KG": (30, 9),
            "materiel": (31, 9),
            "specifique": (32, 9)
        }
    else:
        mapping = {
            "3_6_KG": (23, 9),
            "6_9_KG": (24, 9),
            "9_12_KG": (25, 9),
            "12+_KG": (26, 9),
            "materiel": (27, 9),
            "specifique": (28, 9)
        }

    result = {}

    for key, (r, c) in mapping.items():
        try:
            val = df.iloc[r, c]
            result[key] = float(val) if pd.notna(val) else 0
        except:
            result[key] = 0

    return result


# ----------------------
# MAIN FUNCTION
# ----------------------
def process_files(mec_files, vec_files, engins_file):

    engins_df = pd.read_excel(engins_file)

    # normalisation colonnes engins
    engins_df.columns = engins_df.columns.str.replace("-", "_")

    results = []

    # ---- MEC ----
    for file in mec_files:

        name = file.name.replace(".xlsx", "")
        df = pd.read_excel(file, sheet_name="Résultats", header=10)

        postures = compute_postures_mec(df)

        charges = extract_charges(file, True)

        match = engins_df.loc[engins_df["Poste"] == name]

        if match.empty:
            continue

        engin_row = match.iloc[0]

        results.append({
            "Poste": name,
            "Engin": engin_row["Engin"],
            "engin_debout": engin_row["engin_debout"],
            "engin_frontal": engin_row["engin_frontal"],
            "engin_retract": engin_row["engin_retract"],
            "engin_tous": int(
                engin_row["engin_debout"] == 1 and
                engin_row["engin_frontal"] == 1 and
                engin_row["engin_retract"] == 1
            ),
            "Charge": 0,
            **charges,
            **postures,
            "posture": int(any(postures.values()))
        })

    # ---- VEC ----
    for file in vec_files:

        name = file.name.replace(".xlsx", "")
        df = pd.read_excel(file, sheet_name="Résultats")

        postures = compute_postures_vec(df)

        charges = extract_charges(file, False)

        match = engins_df.loc[engins_df["Poste"] == name]

        if match.empty:
            continue

        engin_row = match.iloc[0]

        results.append({
            "Poste": name,
            "Engin": engin_row["Engin"],
            "engin_debout": engin_row["engin_debout"],
            "engin_frontal": engin_row["engin_frontal"],
            "engin_retract": engin_row["engin_retract"],
            "engin_tous": int(
                engin_row["engin_debout"] == 1 and
                engin_row["engin_frontal"] == 1 and
                engin_row["engin_retract"] == 1
            ),
            "Charge": 0,
            **charges,
            **postures,
            "posture": int(any(postures.values()))
        })

    return pd.DataFrame(results)
